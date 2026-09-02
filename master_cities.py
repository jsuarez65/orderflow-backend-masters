from flask import Flask, request, jsonify
from openpyxl import load_workbook
import psycopg2

app = Flask(__name__)

@app.route('/master/cities', methods=['POST'])
def import_codigos_postales():
    # 1. Validar que se envió un archivo
    if 'file' not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No se envió archivo"}), 400

    # 2. Conexión a la base de datos
    conn = psycopg2.connect(
        host="ep-dry-art-acz5gndj-pooler.sa-east-1.aws.neon.tech",
        port="5432",
        dbname="orderflow",
        user="neondb_owner",
        password="npg_oYRmQ2e0IHaT",
        sslmode="require"
    )
    cursor = conn.cursor()

    # 3. Leer el Excel
    workbook = load_workbook(file)
    sheet = workbook.active

    insertados = 0
    duplicados_ignorados = 0
    saltados = 0

    # 4. Procesar cada fila (min_row=2 para saltar encabezados)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        
        # Saltar filas vacías
        if not row or row[0] is None:
            saltados += 1
            continue

        try:
            # Conversión segura de zipcode (maneja 1001.0 de Excel)
            raw_cp = row[0]
            if isinstance(raw_cp, float):
                codigo_postal = str(int(raw_cp))
            else:
                codigo_postal = str(raw_cp).strip()

            # Validar que no esté vacío
            if not codigo_postal:
                saltados += 1
                continue

            # Nombre de la localidad
            nombre_localidad = str(row[1]).strip().title() if row[1] and str(row[1]).strip() else ""

            # INSERT con ON CONFLICT (la magia de PostgreSQL)
            # Si la combinación (CP + Nombre) ya existe, NO hace nada y no tira error
            cursor.execute("""
                INSERT INTO localidades (codigo_postal, nombre_localidad)
                VALUES (%s, %s)
                ON CONFLICT (codigo_postal, nombre_localidad) DO NOTHING
            """, (codigo_postal, nombre_localidad))

            # cursor.rowcount es 1 si se insertó, 0 si fue ignorado por duplicado
            if cursor.rowcount > 0:
                insertados += 1
            else:
                duplicados_ignorados += 1

        except Exception as e:
            saltados += 1
            print(f"Error en fila {row}: {e}")
            continue

    # 5. Confirmar cambios y cerrar conexión
    conn.commit()
    cursor.close()
    conn.close()

    # 6. Respuesta con estadísticas
    return jsonify({
        "message": "Importación finalizada con éxito",
        "insertados_nuevos": insertados,
        "duplicados_ignorados": duplicados_ignorados,
        "filas_saltadas": saltados,
        "total_procesadas": insertados + duplicados_ignorados + saltados
    })

if __name__ == '__main__':
    app.run(debug=True)