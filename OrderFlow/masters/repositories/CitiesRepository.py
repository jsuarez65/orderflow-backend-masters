from openpyxl import load_workbook
from configuration.DatabaseConfiguration import DatabaseConfiguration
from configuration.LogConfiguration import LogConfiguration

class CitiesRepository:
    def __init__(self):
        self.db = DatabaseConfiguration.getConnection()
        self.log = LogConfiguration.getLogger()

    def importPostalCodes(self, file):
        cursor = self.db.cursor()
        try:
            workbook = load_workbook(file)
            sheet = workbook.active

            inserted = 0
            ignored_duplicates = 0
            skipped = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    skipped += 1
                    continue

                try:
                    raw_cp = row[0]
                    if isinstance(raw_cp, float):
                        postal_code = str(int(raw_cp))
                    else:
                        postal_code = str(raw_cp).strip()

                    if not postal_code:
                        skipped += 1
                        continue

                    city_name = str(row[1]).strip().title() if row[1] and str(row[1]).strip() else ""

                    cursor.execute("""
                        INSERT INTO localidades (codigo_postal, nombre_localidad)
                        VALUES (%s, %s)
                        ON CONFLICT (codigo_postal, nombre_localidad) DO NOTHING
                    """, (postal_code, city_name))

                    if cursor.rowcount > 0:
                        inserted += 1
                    else:
                        ignored_duplicates += 1

                except Exception as e:
                    skipped += 1
                    self.log.error(f"Error in row {row}: {e}")
                    continue

            self.db.commit()
            return {
                "message": "Import finished successfully",
                "inserted_new": inserted,
                "ignored_duplicates": ignored_duplicates,
                "skipped_rows": skipped,
                "total_processed": inserted + ignored_duplicates + skipped
            }
        finally:
            cursor.close()